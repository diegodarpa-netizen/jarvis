from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

VERDE_OSC  = "1A7A4A"
VERDE_MED  = "2ECC71"
VERDE_CLAR = "D5F5E3"
GRIS_OSC   = "2C3E50"
GRIS_CLAR  = "F2F3F4"
BLANCO     = "FFFFFF"
AMARILLO   = "F9E79F"
ROJO_CLAR  = "FADBD8"
AZUL_CLAR  = "D6EAF8"

def hdr(cell, txt, bg=VERDE_OSC, fg=BLANCO, bold=True, size=10, wrap=False):
    cell.value = txt
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)

thin = Side(style='thin', color='AAAAAA')
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_border(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r, c).border = brd

# ── HOJA 1: REGISTRO DIARIO ──
ws1 = wb.active
ws1.title = 'Registro Diario'
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A4'

ws1.merge_cells('A1:N1')
hdr(ws1['A1'], 'REGISTRO DIARIO DE LEADS - Cirugia Estetica', size=13)
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:N2')
hdr(ws1['A2'], 'Completar una fila por cada mensaje recibido. Actualizar estado cuando avance el lead.', bg=GRIS_OSC, size=9)
ws1.row_dimensions[2].height = 16

headers = ['Fecha','Nombre','Telefono','Tratamiento','Fuente','Estado','Turno','Vino?','Cerro?','Monto ARS','Campana','Zona','Msgs Dia','Notas']
widths  = [12,18,14,16,13,16,12,10,10,14,18,12,10,24]
bgs     = [GRIS_CLAR,BLANCO,BLANCO,BLANCO,AZUL_CLAR,AMARILLO,GRIS_CLAR,GRIS_CLAR,GRIS_CLAR,VERDE_CLAR,AZUL_CLAR,AZUL_CLAR,AZUL_CLAR,BLANCO]

for i, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws1.cell(3, i)
    hdr(c, h, bg=VERDE_MED, size=9, wrap=True)
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[3].height = 28

for row in range(4, 64):
    for col in range(1, 15):
        cell = ws1.cell(row, col)
        cell.fill = PatternFill('solid', start_color=bgs[col-1])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(horizontal='center' if col in [1,7,8,9,10,13] else 'left', vertical='center')
    ws1.cell(row, 1).number_format = 'DD/MM/YYYY'
    ws1.cell(row, 7).number_format = 'DD/MM/YYYY'
    ws1.cell(row, 10).number_format = '#,##0'
    ws1.cell(row, 13).font = Font(name='Arial', size=9, color='0000FF', bold=True)
    ws1.row_dimensions[row].height = 16

apply_border(ws1, 3, 63, 1, 14)

ws1['A65'].value = 'ESTADOS:'
ws1['A65'].font = Font(name='Arial', bold=True, size=8)
estados = [('B65','Nuevo',AMARILLO),('C65','Contactado','FFE0B2'),('D65','Turno agendado',VERDE_CLAR),
           ('E65','Consulta realizada',AZUL_CLAR),('F65','Cerrado OK',VERDE_CLAR),
           ('G65','No respondio',ROJO_CLAR),('H65','Perdido',ROJO_CLAR)]
for ref, lbl, col in estados:
    ws1[ref].value = lbl
    ws1[ref].font = Font(name='Arial', size=8, bold=True)
    ws1[ref].fill = PatternFill('solid', start_color=col)
    ws1[ref].alignment = Alignment(horizontal='center')
    ws1[ref].border = brd

ws1['A66'].value = 'FUENTES:'
ws1['A66'].font = Font(name='Arial', bold=True, size=8)
fuentes = [('B66','Meta Ads',AZUL_CLAR),('C66','Organico Instagram','E8F8F5'),
           ('D66','Referido','FEF9E7'),('E66','Google','FFE0B2'),('F66','Otro',GRIS_CLAR)]
for ref, lbl, col in fuentes:
    ws1[ref].value = lbl
    ws1[ref].font = Font(name='Arial', size=8)
    ws1[ref].fill = PatternFill('solid', start_color=col)
    ws1[ref].alignment = Alignment(horizontal='center')
    ws1[ref].border = brd

# ── HOJA 2: RESUMEN MENSUAL ──
ws2 = wb.create_sheet('Resumen Mensual')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:D1')
hdr(ws2['A1'], 'RESUMEN MENSUAL - KPIs de Campana', size=13)
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:D2')
hdr(ws2['A2'], 'Datos tomados del Registro Diario automaticamente', bg=GRIS_OSC, size=9)

for i, h in enumerate(['KPI','Valor','Referencia'], 1):
    hdr(ws2.cell(4, i), h, bg=VERDE_MED, size=9)

kpis = [
    ('Total mensajes recibidos (sumar col Msgs Dia)', '=SUM(B19:B49)', 'Ingresado manualmente en tabla abajo'),
    ('Total leads ingresados',   "=COUNTA('Registro Diario'!B4:B63)", 'Filas con nombre'),
    ('Turnos agendados',         "=COUNTIF('Registro Diario'!F4:F63,\"Turno agendado\")", 'Estado=Turno agendado'),
    ('Consultas realizadas',     "=COUNTIF('Registro Diario'!H4:H63,\"Si\")", 'Col Vino=Si'),
    ('Tratamientos cerrados',    "=COUNTIF('Registro Diario'!I4:I63,\"Si\")", 'Col Cerro=Si'),
    ('No respondieron/perdidos', "=COUNTIF('Registro Diario'!F4:F63,\"No respondio\")+COUNTIF('Registro Diario'!F4:F63,\"Perdido\")", 'Estados negativos'),
    ('Facturacion total ARS',    "=SUM('Registro Diario'!J4:J63)", 'Suma col Monto'),
    ('Ticket promedio ARS',      '=IFERROR(B11/B9,0)', 'Facturacion/Cierres'),
    ('% Lead a turno',           '=IFERROR(B7/B6,0)', 'Turnos/Leads'),
    ('% Turno a cierre',         '=IFERROR(B9/B7,0)', 'Cierres/Turnos'),
    ('% Lead a cierre',          '=IFERROR(B9/B6,0)', 'Cierres/Leads'),
]

for i, (label, formula, nota) in enumerate(kpis, 5):
    ws2.cell(i, 1).value = label
    ws2.cell(i, 1).font = Font(name='Arial', bold=True, size=9)
    ws2.cell(i, 1).fill = PatternFill('solid', start_color=GRIS_CLAR)
    ws2.cell(i, 2).value = formula
    ws2.cell(i, 2).font = Font(name='Arial', size=10, bold=True, color='0000FF')
    ws2.cell(i, 2).fill = PatternFill('solid', start_color=VERDE_CLAR)
    ws2.cell(i, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(i, 3).value = nota
    ws2.cell(i, 3).font = Font(name='Arial', size=8, italic=True, color='666666')
    ws2.row_dimensions[i].height = 18

for r in [13,14,15]:
    ws2.cell(r,2).number_format = '0.0%'
for r in [11,12]:
    ws2.cell(r,2).number_format = '#,##0'

apply_border(ws2, 4, 15, 1, 3)

ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 38

ws2.merge_cells('A17:C17')
hdr(ws2['A17'], 'MENSAJES POR DIA - Ingresar manualmente cada dia', bg=VERDE_OSC, size=10)
for i, h in enumerate(['Fecha','Mensajes del dia','Leads ingresados'], 1):
    hdr(ws2.cell(18, i), h, bg=VERDE_MED, size=9)

for r in range(19, 50):
    ws2.cell(r,1).number_format = 'DD/MM/YYYY'
    ws2.cell(r,1).font = Font(name='Arial', size=9, color='0000FF', bold=True)
    ws2.cell(r,1).fill = PatternFill('solid', start_color=GRIS_CLAR)
    ws2.cell(r,2).font = Font(name='Arial', size=10, color='0000FF', bold=True)
    ws2.cell(r,2).fill = PatternFill('solid', start_color=AZUL_CLAR)
    ws2.cell(r,2).alignment = Alignment(horizontal='center')
    ws2.cell(r,3).font = Font(name='Arial', size=10, color='0000FF', bold=True)
    ws2.cell(r,3).fill = PatternFill('solid', start_color=VERDE_CLAR)
    ws2.cell(r,3).alignment = Alignment(horizontal='center')
    ws2.row_dimensions[r].height = 15

ws2['A50'].value = 'TOTAL DEL MES'
ws2['A50'].font = Font(name='Arial', bold=True, size=10)
ws2['A50'].fill = PatternFill('solid', start_color=AMARILLO)
ws2['B50'].value = '=SUM(B19:B49)'
ws2['B50'].font = Font(name='Arial', bold=True, size=12, color=VERDE_OSC)
ws2['B50'].fill = PatternFill('solid', start_color=AMARILLO)
ws2['B50'].alignment = Alignment(horizontal='center')
ws2['C50'].value = '=SUM(C19:C49)'
ws2['C50'].font = Font(name='Arial', bold=True, size=12, color=VERDE_OSC)
ws2['C50'].fill = PatternFill('solid', start_color=AMARILLO)
ws2['C50'].alignment = Alignment(horizontal='center')

apply_border(ws2, 17, 50, 1, 3)

# ── HOJA 3: POR CAMPANA ──
ws3 = wb.create_sheet('Por Campana')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:G1')
hdr(ws3['A1'], 'ANALISIS POR CAMPANA - ROI y Conversion', size=13)
ws3.row_dimensions[1].height = 30

for i, h in enumerate(['Campana','Gasto Meta ARS/mes','Leads','Cierres','Facturacion ARS','Costo x Lead ARS','ROI'], 1):
    hdr(ws3.cell(3, i), h, bg=VERDE_MED, size=9, wrap=True)
ws3.row_dimensions[3].height = 28

campanas = [
    'Implantes Mamarios - CABA/San Isidro',
    'Implantes Mamarios - GBA Sur/Pilar',
    'Medicina Estetica - CABA',
    'Cirugia General',
    'Uruguay - Montevideo',
]
for i, nombre in enumerate(campanas, 4):
    ws3.cell(i,1).value = nombre
    ws3.cell(i,1).font = Font(name='Arial', bold=True, size=9)
    ws3.cell(i,1).fill = PatternFill('solid', start_color=GRIS_CLAR)
    for c in [2,3,4,5]:
        ws3.cell(i,c).font = Font(name='Arial', size=10, bold=True, color='0000FF')
        ws3.cell(i,c).fill = PatternFill('solid', start_color=AZUL_CLAR)
        ws3.cell(i,c).alignment = Alignment(horizontal='center')
    ws3.cell(i,6).value = f'=IFERROR(B{i}/C{i},0)'
    ws3.cell(i,6).number_format = '#,##0'
    ws3.cell(i,6).font = Font(name='Arial', size=9)
    ws3.cell(i,6).fill = PatternFill('solid', start_color=AMARILLO)
    ws3.cell(i,6).alignment = Alignment(horizontal='center')
    ws3.cell(i,7).value = f'=IFERROR((E{i}-B{i})/B{i},0)'
    ws3.cell(i,7).number_format = '0.0%'
    ws3.cell(i,7).font = Font(name='Arial', bold=True, size=9)
    ws3.cell(i,7).fill = PatternFill('solid', start_color=VERDE_CLAR)
    ws3.cell(i,7).alignment = Alignment(horizontal='center')
    ws3.row_dimensions[i].height = 18

r = len(campanas) + 4
ws3.cell(r,1).value = 'TOTAL'
ws3.cell(r,1).font = Font(name='Arial', bold=True, size=10)
ws3.cell(r,1).fill = PatternFill('solid', start_color=AMARILLO)
for ci in [2,3,4,5]:
    col = get_column_letter(ci)
    ws3.cell(r,ci).value = f'=SUM({col}4:{col}{r-1})'
    ws3.cell(r,ci).font = Font(name='Arial', bold=True, size=10, color=VERDE_OSC)
    ws3.cell(r,ci).fill = PatternFill('solid', start_color=AMARILLO)
    ws3.cell(r,ci).alignment = Alignment(horizontal='center')
ws3.cell(r,6).value = f'=IFERROR(B{r}/C{r},0)'
ws3.cell(r,6).number_format = '#,##0'
ws3.cell(r,6).font = Font(name='Arial', bold=True)
ws3.cell(r,6).fill = PatternFill('solid', start_color=AMARILLO)
ws3.cell(r,6).alignment = Alignment(horizontal='center')
ws3.cell(r,7).value = f'=IFERROR((E{r}-B{r})/B{r},0)'
ws3.cell(r,7).number_format = '0.0%'
ws3.cell(r,7).font = Font(name='Arial', bold=True)
ws3.cell(r,7).fill = PatternFill('solid', start_color=AMARILLO)
ws3.cell(r,7).alignment = Alignment(horizontal='center')

apply_border(ws3, 3, r, 1, 7)
ws3.column_dimensions['A'].width = 38
for col in ['B','C','D','E','F','G']:
    ws3.column_dimensions[col].width = 16

ws3.cell(r+2, 1).value = 'Completar columnas B,C,D,E con datos reales. Las formulas calculan Costo x Lead y ROI automaticamente.'
ws3.cell(r+2, 1).font = Font(name='Arial', size=8, italic=True, color='666666')
ws3.merge_cells(f'A{r+2}:G{r+2}')

output = r'C:\Users\prometech\Desktop\Jarvis\marketing\Control_Leads_Campanas.xlsx'
wb.save(output)
print('GUARDADO:', output)
